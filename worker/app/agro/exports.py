"""Independent Excel exports for the agrometeorological bulletin."""

from __future__ import annotations

from io import BytesIO
from typing import Iterable

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .calculations import grouped_stations, resolve_etp_station
from .models import DailyAgro, DailyRain, EditableDecadeValues, RainfallNormal, Station

MONTHS = ("JANVIER", "FEVRIER", "MARS", "AVRIL", "MAI", "JUIN", "JUILLET", "AOUT", "SEPTEMBRE", "OCTOBRE", "NOVEMBRE", "DECEMBRE")
NETWORK_HEADERS = ["STATIONS", "LONGITUDE", "LATITUDE", "Nbre jours pluie > 00mm", "Nbre jours pluie > 20mm", "Sur la décade en cours", "Ecart à la normale", "% de la normale", "Depuis début année civile", "Ecart à la normale", "Depuis début Saison des pluies", "Ecart à la normale", "Bilan hydrique"]


def _download(workbook: openpyxl.Workbook, filename: str) -> tuple[BytesIO, str]:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream, filename


def _style_table(sheet: openpyxl.worksheet.worksheet.Worksheet, header_row: int, widths: list[int]) -> None:
    border = Border(*(Side(style="thin", color="9BB7A2") for _ in range(4)))
    for row in sheet.iter_rows(min_row=header_row, max_row=sheet.max_row, min_col=1, max_col=len(widths)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="196B3A")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    sheet.freeze_panes = None
    sheet.sheet_view.showGridLines = True
    sheet.sheet_view.zoomScale = 90


def build_network_export(year: int, month: int, decade: int, stations: Iterable[Station], summaries: dict[str, object]) -> tuple[BytesIO, str]:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Réseau pluviométrique"
    groups = (("TABLEAU 1", ("Alibori", "Atacora", "Borgou", "Donga")), ("TABLEAU 2", ("Collines", "Couffo", "Mono", "Zou")), ("TABLEAU 3", ("Atlantique", "Littoral", "Oueme", "Plateau")))
    stations = list(stations)
    for table_number, departments in groups:
        if sheet.max_row > 1:
            sheet.append([])
        start = sheet.max_row + 1
        sheet.merge_cells(start_row=start, start_column=1, end_row=start, end_column=12)
        sheet.cell(start, 1, f"ANNEE : {year} | MOIS : {MONTHS[month - 1]} | DECADE : {decade} | {table_number}")
        sheet.cell(start, 1).font = Font(bold=True, size=13, color="FFFFFF")
        sheet.cell(start, 1).fill = PatternFill("solid", fgColor="0D472B")
        sheet.cell(start, 1).alignment = Alignment(horizontal="center")
        sheet.merge_cells(start_row=start + 1, start_column=1, end_row=start + 1, end_column=12)
        sheet.cell(start + 1, 1, "RESEAU PLUVIOMETRIQUE - DEPARTEMENTS : " + ", ".join(departments))
        sheet.cell(start + 1, 1).font = Font(bold=True)
        sheet.append(NETWORK_HEADERS)
        header_row = sheet.max_row
        for department in departments:
            members = [station for station in stations if station.department == department]
            if not members:
                continue
            department_row = sheet.max_row + 1
            sheet.append([department])
            for cell in sheet[department_row][:12]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4E8B5D")
            for station in members:
                summary = summaries.get(station.id, {})
                row_number = sheet.max_row + 1
                normal = summary.get("normal_decade")
                etp = summary.get("etp")
                daily = list(summary.get("daily_values", []))[:10]
                daily += [None] * (10 - len(daily))
                sheet.append([station.name, summary.get("longitude", station.longitude), summary.get("latitude", station.latitude), f'=COUNTIF(N{row_number}:W{row_number},">0")', f'=COUNTIF(N{row_number}:W{row_number},">20")', f'=SUM(N{row_number}:W{row_number})', f'=F{row_number}-{normal}' if isinstance(normal, (int, float)) else summary.get("decade_deviation"), f'=IFERROR(F{row_number}/{normal},"")' if isinstance(normal, (int, float)) and normal else summary.get("normal_percentage"), summary.get("year_total"), summary.get("year_deviation"), summary.get("season_total"), summary.get("season_deviation"), f'=F{row_number}-{etp}' if isinstance(etp, (int, float)) else summary.get("water_balance")] + daily)
        _style_table(sheet, header_row, [24, 13, 13, 14, 14, 18, 18, 18, 20, 18, 24, 18, 16] + [10] * 10)
    return _download(workbook, f"DONNEES_PLUVIOMETRIQUES_{year}_{month:02d}_D{decade}.xlsx")


def build_climate_export(year: int, month: int, decade: int, stations: Iterable[Station], climate: dict[str, dict[str, object]]) -> tuple[BytesIO, str]:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Données climatiques"
    sheet.merge_cells("A1:I1")
    sheet["A1"] = "V-a - DONNEES CLIMATIQUES COMPLEMENTAIRES"
    sheet["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0D472B")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:I2")
    sheet["A2"] = f"Période : {decade}ère décade de {MONTHS[month - 1].title()} {year}"
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.append([])
    sheet.append(["STATIONS", "Durée Insolation h./10", "Fraction Insolation %", "Rayonn. Global j/cm²", "Vent moyen", "Vent maxi.", "EVAPO. Bac", "ETP Penman", "Bilan hydrique potentiel"])
    for station in stations:
        values = climate.get(station.id, {})
        sheet.append([station.name] + [values.get(key) for key in ("sunshine_total", "insolation_fraction", "global_radiation", "wind_mean", "wind_max", "pan_evaporation", "etp", "water_balance")])
    _style_table(sheet, 4, [22, 22, 22, 24, 16, 16, 16, 16, 25])
    sheet.auto_filter.ref = f"A4:I{sheet.max_row}"
    normals_sheet = workbook.create_sheet("Tableau IV")
    normals_sheet.merge_cells("A1:J1")
    normals_sheet["A1"] = "TABLEAU IV - DONNEES CLIMATIQUES (Moyennes sur décade)"
    normals_sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    normals_sheet["A1"].fill = PatternFill("solid", fgColor="0D472B")
    normals_sheet.append(["STATIONS", "Tmin", "Tmax", "Tmoy", "+10cm", "+50cm", "Hum. min", "Hum. max", "Hum. moy", "Tension Vapeur", "Déficit"])
    for station in stations:
        values = climate.get(station.id, {})
        normal = values.get("normal") or {}
        current = [values.get(key) for key in ("tmin", "tmax", "tmean", "soil10", "soil50", "humidity_min", "humidity_max", "humidity_mean", "vapor_pressure", "deficit")]
        normals_sheet.append([station.name] + [value if value is not None else None for value in current])
        normals_sheet.append(["Ecart/Normale"] + [current[index] - normal[key] if isinstance(current[index], (int, float)) and isinstance(normal.get(key), (int, float)) else None for index, key in enumerate(("tmin", "tmax", "tmean", "soil10", "soil50", "hmin", "hmax", "hmean", "vapor_pressure", "deficit"))])
    _style_table(normals_sheet, 2, [22, 14, 14, 14, 14, 14, 14, 14, 14, 18, 14])
    return _download(workbook, f"DONNEES_CLIMATIQUES_{year}_{month:02d}_D{decade}.xlsx")


def build_observations_export(year: int, month: int, decade: int, station: Station, rows: list[dict[str, object]]) -> tuple[BytesIO, str]:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Renseignements agro"
    sheet.merge_cells("A1:J1")
    sheet["A1"] = "RENSEIGNEMENTS AGROMETEOROLOGIQUES"
    sheet["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0D472B")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:J2")
    sheet["A2"] = f"Station : {station.name} | Période : {decade}ère décade de {MONTHS[month - 1]} {year}"
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.append([])
    headers = ["Jour", "Pluie", "Insolation", "Tmin", "Tmax", "T moy", "Hum. min", "Hum. max", "Hum. moy", "Évapo. bac"]
    sheet.append(headers)
    data_start = sheet.max_row + 1
    ordered = {int(row.get("jour", 0)): row for row in rows}
    for day in range(1, 11):
        row = ordered.get(day, {})
        tmin, tmax = row.get("temp_min"), row.get("temp_max")
        hmin, hmax = row.get("humidite_min"), row.get("humidite_max")
        row_number = sheet.max_row + 1
        sheet.append([day, row.get("pluie"), row.get("insolation"), tmin, tmax, f"=IF(COUNT(D{row_number}:E{row_number})=2,AVERAGE(D{row_number}:E{row_number}),\"\")", hmin, hmax, f"=IF(COUNT(G{row_number}:H{row_number})=2,0.6*G{row_number}+0.4*H{row_number},\"\")", row.get("evapo_bac_a")])
    data_end = sheet.max_row
    sheet.append(["Total", f"=SUM(B{data_start}:B{data_end})", f"=SUM(C{data_start}:C{data_end})", "", "", "", "", "", "", f"=SUM(J{data_start}:J{data_end})"])
    sheet.append(["Moyenne", f"=AVERAGE(B{data_start}:B{data_end})", f"=AVERAGE(C{data_start}:C{data_end})", f"=AVERAGE(D{data_start}:D{data_end})", f"=AVERAGE(E{data_start}:E{data_end})", f"=AVERAGE(F{data_start}:F{data_end})", f"=AVERAGE(G{data_start}:G{data_end})", f"=AVERAGE(H{data_start}:H{data_end})", f"=AVERAGE(I{data_start}:I{data_end})", f"=AVERAGE(J{data_start}:J{data_end})"])
    _style_table(sheet, 4, [12, 14, 16, 12, 12, 12, 14, 14, 14, 16])
    for row in (sheet.max_row - 1, sheet.max_row):
        for cell in sheet[row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4E8B5D")
    sheet.auto_filter.ref = f"A4:J{sheet.max_row}"
    return _download(workbook, f"RENSEIGNEMENTS_AGRO_{station.id}_{year}_{month:02d}_D{decade}.xlsx")
    
    sheet.append([])
    sheet.append(["* L'humidité moyenne (Umoy) est calculée à partir de la température moyenne."])
    sheet.append(["* Déficit de saturation = ew - tension de vapeur moyenne."])
    sheet.append(["* Les données manquantes sont codées par -."])
    for row in (sheet[2],):
        for cell in row: cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="D9EAD3")
    return _download(workbook, f"DONNEES_CLIMATIQUES-{decade}_{MONTHS[month - 1]}_{year}.xlsx")
