"""Rainfall workbook ingestion and RESA-01 calculations."""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any

import xlrd
import openpyxl


@dataclass(frozen=True)
class RainfallImport:
    year: int
    month: int
    decade: int
    source: str
    rows: list[dict[str, Any]]


MONTHS = {"JANVIER": 1, "FEVRIER": 2, "FÉVRIER": 2, "MARS": 3, "AVRIL": 4, "MAI": 5, "JUIN": 6, "JUILLET": 7, "AOUT": 8, "AOÛT": 8, "SEPTEMBRE": 9, "OCTOBRE": 10, "NOVEMBRE": 11, "DECEMBRE": 12, "DÉCEMBRE": 12}


def _number(value: object) -> float | None:
    if value in (None, "", '"'):
        return None
    if isinstance(value, str) and value.strip().upper() in {"T", "TR"}:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _metadata(sheet: Any) -> tuple[int, int, int]:
    values = [str(sheet.cell_value(r, c)).strip() for r in range(min(sheet.nrows, 12)) for c in range(sheet.ncols)]
    year = next((int(float(v)) for v in values if re.fullmatch(r"20\d{2}(?:\.0)?", v)), 0)
    month_name = next((v.upper() for v in values if v.upper() in MONTHS), "")
    decade_map = {"I": 1, "II": 2, "III": 3, "1ERE": 1, "1ÈRE": 1, "2EME": 2, "2ÈME": 2, "3EME": 3, "3ÈME": 3}
    decade = next((decade_map[v.upper()] for v in values if v.upper() in decade_map), 0)
    if not year or not month_name or not decade:
        raise ValueError("Métadonnées année/mois/décade introuvables")
    return year, MONTHS[month_name], decade


def _parse_decades_sheet(sheet: Any, cell: Any, metadata: tuple[int, int, int] | None = None) -> RainfallImport:
    rows_count = sheet.nrows if hasattr(sheet, "nrows") else sheet.max_row
    cols_count = sheet.ncols if hasattr(sheet, "ncols") else sheet.max_column
    value = lambda r, c: sheet.cell_value(r, c) if hasattr(sheet, "cell_value") else sheet.cell(r + 1, c + 1).value
    class Adapter:
        nrows = rows_count
        ncols = cols_count
        def cell_value(self, r: int, c: int) -> Any:
            return value(r, c)
        def cell(self, r: int, c: int) -> Any:
            return type("Cell", (), {"value": value(r, c)})()
    adapted = Adapter()
    year, month, decade = metadata or _metadata(adapted)
    rows: list[dict[str, Any]] = []
    department = ""
    for r in range(rows_count):
        values = [value(r, c) for c in range(cols_count)]
        text = " ".join(str(v).strip() for v in values[:3] if v not in ("", None)).strip()
        normalized = re.sub(r"\s+", " ", text.upper()).replace("É", "E")
        if normalized in {"ATLANTIQUE-LITTORAL", "ATACORA-DONGA", "BORGOU-ALIBORI", "MONO-COUFFO", "OUEME-PLATEAU", "ZOU-COLLINE"}:
            department = text
            continue
        localite = str(values[1]).strip() if len(values) > 1 and values[1] not in ("", None) else ""
        if not department and localite:
            department = str(values[0]).strip()
        if not department or not localite or localite.upper() in {"LOCALITES", "DEPARTEMENTS"}:
            continue
        daily = [_number(values[c]) for c in range(2, min(13, len(values)))]
        if any(value is not None for value in daily):
            rows.append({"department": department, "station": localite, "daily": daily})
    return RainfallImport(year, month, decade, "decades", rows)


def parse_decades_xls(content: bytes) -> RainfallImport:
    book = xlrd.open_workbook(file_contents=content)
    sheet = next((s for s in book.sheets() if s.nrows and any("HAUTEURS JOURNALIERES" in str(s.cell_value(r, c)).upper() for r in range(min(s.nrows, 10)) for c in range(s.ncols))), None)
    if sheet is None:
        raise ValueError("Feuille DECADES introuvable")
    return _parse_decades_sheet(sheet, None)


def parse_decades_xlsx(content: bytes, metadata: tuple[int, int, int] | None = None) -> RainfallImport:
    book = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    sheet = next((s for s in book.worksheets if s.max_row and any("HAUTEURS JOURNALIERES" in str(s.cell(r, c).value).upper() for r in range(1, min(s.max_row, 10) + 1) for c in range(1, s.max_column + 1))), None)
    if sheet is None:
        raise ValueError("Feuille DECADES introuvable")
    if metadata is None:
        metadata = _metadata_from_text(" ".join(str(sheet.cell(r, c).value) for r in range(1, min(sheet.max_row, 8) + 1) for c in range(1, sheet.max_column + 1)))
    return _parse_decades_sheet(sheet, None, metadata)


def _metadata_from_text(text: str) -> tuple[int, int, int]:
    upper = text.upper()
    year_match = re.search(r"\b(20\d{2})\b", upper)
    month_match = next((name for name in MONTHS if name in upper), None)
    decade_match = re.search(r"(?:1(?:ERE|ÈRE)|2(?:EME|ÈME)|3(?:EME|ÈME))", upper)
    if not year_match or not month_match or not decade_match:
        raise ValueError("Métadonnées année/mois/décade introuvables; renseignez la période")
    decade = int(decade_match.group()[0])
    return int(year_match.group(1)), MONTHS[month_match], decade


def parse_agro_xls(content: bytes) -> RainfallImport:
    book = xlrd.open_workbook(file_contents=content)
    sheet = next((s for s in book.sheets() if s.nrows and any("RENSEIGNEMENTS AGROMETEOROLOGIQUES" in str(s.cell_value(r, c)).upper() for r in range(min(s.nrows, 3)) for c in range(s.ncols))), None)
    if sheet is None:
        raise ValueError("Feuille opérationnelle Renseignements Agro introuvable")
    year, month, decade = _metadata(sheet)
    rows: list[dict[str, Any]] = []
    station = ""
    for r in range(sheet.nrows):
        first = str(sheet.cell_value(r, 0)).strip() if sheet.ncols else ""
        station_text = next((str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols) if re.match(r"STATION\s*:", str(sheet.cell_value(r, c)).strip(), re.I)), "")
        station_match = re.match(r"STATION\s*:\s*(.+)", station_text or first, re.I)
        if station_match:
            station = station_match.group(1).strip()
            continue
        day = _number(sheet.cell_value(r, 0)) if sheet.ncols else None
        if day is None and first.upper().startswith("STATION"):
            station = first.split(":", 1)[-1].strip()
            continue
        if not station or day is None or not 1 <= day <= 31:
            continue
        rows.append({
            "station": station,
            "day": int(day),
            "rain": _number(sheet.cell_value(r, 1)),
            "tmin": _number(sheet.cell_value(r, 2)),
            "tmax": _number(sheet.cell_value(r, 3)),
            "tmean": _number(sheet.cell_value(r, 4)),
            "soil10": _number(sheet.cell_value(r, 5)),
            "soil50": _number(sheet.cell_value(r, 6)),
            "windMean": _number(sheet.cell_value(r, 7)),
            "windMax": _number(sheet.cell_value(r, 8)),
            "sunshine": _number(sheet.cell_value(r, 9)),
            "humidityMin": _number(sheet.cell_value(r, 10)),
            "humidityMax": _number(sheet.cell_value(r, 11)),
            "humidityMean": _number(sheet.cell_value(r, 12)),
            "vaporPressure": _number(sheet.cell_value(r, 13)),
            "evapPan": _number(sheet.cell_value(r, 14)),
        })
    if not rows:
        raise ValueError("Aucune observation journalière synoptique trouvée")
    return RainfallImport(year, month, decade, "agro", rows)


def parse_rainfall_normals_xls(content: bytes) -> list[dict[str, Any]]:
    sheet = xlrd.open_workbook(file_contents=content).sheet_by_name("Normales")
    result: list[dict[str, Any]] = []
    for col in range(sheet.ncols):
        station = str(sheet.cell_value(1, col)).strip()
        if not station or station.upper() == "DECADES":
            continue
        if col + 2 >= sheet.ncols or str(sheet.cell_value(1, col + 1)).strip().upper() != "CUMA":
            continue
        for row in range(2, sheet.nrows):
            code = str(sheet.cell_value(row, col - 1 if col else col)).strip().lower()
            if not re.fullmatch(r"(?:j|f|m|a|ma|ju|ju|au|s|o|n|d|a)\d|[a-z]+\d", code):
                continue
            result.append({"station": station, "decadeCode": code, "cuma": _number(sheet.cell_value(row, col + 1)), "cums": _number(sheet.cell_value(row, col + 2)), "source": "rainfall"})
    return result


def parse_agro_normals_xls(content: bytes) -> list[dict[str, Any]]:
    sheet = xlrd.open_workbook(file_contents=content).sheet_by_name("Normales")
    result: list[dict[str, Any]] = []
    station = ""
    for row in range(sheet.nrows):
        first = str(sheet.cell_value(row, 0)).strip()
        if first and not re.match(r"(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\d_m$", first, re.I):
            station = first
            continue
        if not station or not first:
            continue
        result.append({"station": station, "decadeCode": first.lower(), "tmin": _number(sheet.cell_value(row, 1)), "tmax": _number(sheet.cell_value(row, 2)), "tmean": _number(sheet.cell_value(row, 3)), "humidityMin": _number(sheet.cell_value(row, 4)), "humidityMax": _number(sheet.cell_value(row, 5)), "humidityMean": _number(sheet.cell_value(row, 6)), "sunshine": _number(sheet.cell_value(row, 7)), "source": "agro"})
    return result


def compute_resa_row(row: dict[str, Any], normal_cuma: float | None = None, etp: float | None = None) -> dict[str, Any]:
    daily = [value for value in row.get("daily", []) if value is not None]
    total = sum(daily)
    return {**row, "nbRainDays": sum(value >= 1 for value in daily), "nbOver20": sum(value > 20 for value in daily), "maxDaily": max(daily, default=None), "decadeTotal": total, "decadeDeviation": total - normal_cuma if normal_cuma is not None else None, "waterBalance": total - etp if etp is not None else None}
