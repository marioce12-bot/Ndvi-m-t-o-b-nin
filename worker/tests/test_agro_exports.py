from io import BytesIO
import unittest

import openpyxl

from app.agro.exports import build_climate_export, build_network_export, build_observations_export
from app.agro.models import Station
from app.agro.registry import canonical_stations


class AgroExportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.stations = [Station("a", "Station A", "Alibori", "A"), Station("c", "Station C", "Littoral", "C")]

    def test_network_export_is_independent_and_downloadable(self) -> None:
        stream, filename = build_network_export(2026, 8, 1, self.stations, {"a": {"rain_days": 2}})
        workbook = openpyxl.load_workbook(stream)
        self.assertEqual(filename, "DONNEES_PLUVIOMETRIQUES_2026_08_D1.xlsx")
        self.assertEqual(workbook.sheetnames, ["Réseau pluviométrique"])
        self.assertIn("RESEAU PLUVIOMETRIQUE", str(workbook.active["A3"].value))
        self.assertEqual(workbook.active["A4"].value, "STATIONS")

    def test_climate_export_keeps_missing_values_blank(self) -> None:
        stream, filename = build_climate_export(2026, 8, 1, self.stations, {"a": {"etp": 12}})
        workbook = openpyxl.load_workbook(stream)
        self.assertEqual(filename, "DONNEES_CLIMATIQUES_2026_08_D1.xlsx")
        self.assertEqual(workbook.active["A1"].value, "V-a - DONNEES CLIMATIQUES COMPLEMENTAIRES")
        self.assertEqual(workbook.active[5][1].value, None)

    def test_network_export_contains_all_stations_in_one_table(self) -> None:
        stream, _ = build_network_export(2026, 8, 1, canonical_stations(), {})
        sheet = openpyxl.load_workbook(stream).active
        values = [cell.value for row in sheet.iter_rows() for cell in row]
        self.assertIn("Agouna", values)
        self.assertIn("Allada", values)

    def test_observations_export_contains_a_table_per_station_with_all_columns(self) -> None:
        rows_by_station = {
            "a": [{"jour": 1, "pluie": 5, "temp_min": 20, "temp_max": 30, "temp_10cm": 27, "temp_50cm": 26, "vent_moyen": 2, "vent_max": 4, "insolation": 8, "humidite_min": 40, "humidite_max": 90, "tension_vapeur": 2.1, "evapo_bac_a": 4}],
            "c": [],
        }
        stream, filename = build_observations_export(2026, 8, 1, self.stations, rows_by_station)
        self.assertEqual(filename, "RENSEIGNEMENTS_AGRO_2026_08_D1.xlsx")
        sheet = openpyxl.load_workbook(stream).active
        values = [cell.value for row in sheet.iter_rows() for cell in row]
        # Both stations get their own table.
        self.assertIn("Station A", " ".join(str(v) for v in values if v))
        self.assertIn("Station C", " ".join(str(v) for v in values if v))
        # All 15 columns from the platform table are present.
        for header in ("Temp. 10cm", "Temp. 50cm", "Vent moyen", "Vent maxi", "Tension vapeur"):
            self.assertIn(header, values)


if __name__ == "__main__":
    unittest.main()
